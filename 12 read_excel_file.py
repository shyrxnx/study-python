import openpyxl
import os

os.chdir(r"C:\Others\Folders\Al Sweigart\Materials\automate_online-materials")
file_name = os.path.basename(r"C:\Others\Folders\Al Sweigart\Materials\automate_online-materials\example.xlsx")

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

sheet = workbook["First"]
print(type(sheet))

# # Review the names of all worksheets of the workbook
# print(workbook.sheetnames)
#
# # Loop through the sheets
# print(f"The sheets in the {file_name} are:")
# for sheet in workbook:
#     print(sheet.title)

# Access A1 cell
first_cell = sheet['A1']
print(first_cell.value)

# Access C3 cell
c_cell = sheet['C3']
print(c_cell.value)

# Provides access to cells using row and column notation
row_col = sheet.cell(row=1, column=1)
print(row_col.value)

# Loops through the values of row = i and column = 1
print("\n")
print("The dates are: ")
for content in range(1, 8):
    print(f"{content}. {sheet.cell(row=content, column=1).value}")

# Loops through the values of row = i and column = 2
print("\n")
print("The fruits are: ")
for content in range(1, 8):
    print(f"{content}. {sheet.cell(row=content, column=2).value}")

# Loops through the values of row = i and column = 3
print("\n")
print("The integers are: ")
for content in range(1, 8):
    print(f"{content}. {sheet.cell(row=content, column=3).value}")
