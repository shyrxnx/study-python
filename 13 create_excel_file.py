from openpyxl import Workbook
import os

os.chdir(r"C:\Others\Python Documents")

# Creates a new workbook
new_file = Workbook()

# Removed the initial sheet
new_file.remove(new_file['Sheet'])

# Creates a new sheets
sheet_one = new_file.create_sheet("First Sheet")
sheet_two = new_file.create_sheet("Second Sheet")
sheet_three = new_file.create_sheet("Third Sheet")

# Renames the third sheet
sheet_three.title = "THE THIRD SHEET"

# Print all sheets
print(new_file.sheetnames)

# Add data to the cells
sheet_one['A1'] = "Dog"
sheet_one['A2'] = "Cat"
sheet_one['B1'] = 1
sheet_one['B2'] = 3

sheet_two['A1'] = "Snake"
sheet_two['B1'] = 1

sheet_three['A1'] = "Ladders"
sheet_three['B1'] = 3

# Save file in the directory -> C:\Others\Python Documents
new_file.save(r'first_excel_file_openpyxl.xlsx')
